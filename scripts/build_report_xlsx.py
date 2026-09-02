"""
Build the Eye Tracking Report workbook from outputs/report-data.json.

Mirrors the supplied template: a Summary sheet of participants against
real-time and offline accuracy, then one sheet per participant listing each
test with a picture of the tracking and a link to the file behind it.

Run scripts/export-report-data.ts first, or use `npm run report:xlsx`, which
chains the two.

Usage: python3 scripts/build_report_xlsx.py [--in outputs/report-data.json]
                                            [--out outputs] [--no-images]
"""
from __future__ import annotations

import argparse
import json
import os
from datetime import datetime

import matplotlib

matplotlib.use("Agg")
import matplotlib.pyplot as plt
from matplotlib.collections import LineCollection
import numpy as np
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# Template geometry, kept as supplied.
SUMMARY_WIDTHS = {"A": 10.7, "B": 20.6, "C": 18.1}
PARTICIPANT_WIDTHS = {"A": 13.9, "B": 61.4, "C": 33.2}

# Column B is 61.4 characters wide, which is about 430 px, so the picture is
# drawn to sit inside it rather than being scaled by Excel.
IMG_W_PX = 420
DPI = 100

HEADER_FILL = PatternFill("solid", fgColor="1E293B")
HEADER_FONT = Font(color="FFFFFF", bold=True)
NOTE_FONT = Font(color="64748B", italic=True, size=9)
DEG_FMT = '0.00"°"'


def render_gaze(test: dict, path: str) -> bool:
    """Draw one test's gaze path. Returns False when there was nothing to draw."""
    pts = test.get("gazePath") or []
    vw = test.get("viewportWidth") or 1536
    vh = test.get("viewportHeight") or 864
    aspect = vh / vw
    fig_w = IMG_W_PX / DPI
    fig, ax = plt.subplots(figsize=(fig_w, fig_w * aspect), dpi=DPI)

    ax.set_xticks([])
    ax.set_yticks([])
    for spine in ax.spines.values():
        spine.set_visible(False)
    ax.set_facecolor("white")

    # The screen itself, drawn rather than assumed: the regressor can predict a
    # point beyond the display, and clipping to the viewport would hide exactly
    # the samples worth seeing. So the frame is a rectangle and the axes stretch
    # to whatever the data did.
    ax.add_patch(
        plt.Rectangle((0, 0), vw, vh, fill=True, facecolor="#F8FAFC",
                      edgecolor="#94A3B8", linewidth=0.8, zorder=0)
    )

    if not pts:
        ax.set_xlim(0, vw)
        ax.set_ylim(vh, 0)
        ax.text(
            vw / 2, vh / 2,
            "No gaze path recorded",
            ha="center", va="center", color="#94A3B8", fontsize=9,
        )
    else:
        x = np.array([p["x"] for p in pts], dtype=float)
        y = np.array([p["y"] for p in pts], dtype=float)
        t = np.array([p.get("t", i) for i, p in enumerate(pts)], dtype=float)

        # Colour the trace by time so the direction of travel is readable
        # without arrows, which clutter at this size.
        segs = np.stack([np.column_stack([x[:-1], y[:-1]]), np.column_stack([x[1:], y[1:]])], axis=1)
        norm = plt.Normalize(t.min(), t.max()) if t.max() > t.min() else plt.Normalize(0, 1)
        lc = LineCollection(segs, cmap="viridis", norm=norm, linewidth=1.0, alpha=0.85)
        lc.set_array(t[:-1])
        ax.add_collection(lc)

        ax.scatter([x[0]], [y[0]], s=26, c="#16A34A", zorder=3, edgecolors="white", linewidths=0.8)
        ax.scatter([x[-1]], [y[-1]], s=26, c="#DC2626", zorder=3, edgecolors="white", linewidths=0.8)

        pad = 0.03 * max(vw, vh)
        ax.set_xlim(min(0, x.min()) - pad, max(vw, x.max()) + pad)
        ax.set_ylim(max(vh, y.max()) + pad, min(0, y.min()) - pad)  # y runs downwards

        off = int(((x < 0) | (x > vw) | (y < 0) | (y > vh)).sum())
        caption = f"{len(pts)} samples · {vw}×{vh} · green = start, red = end"
        if off:
            caption += f" · {off} outside the screen"
        # Below the axes, not inside them: once the limits stretch past the
        # screen the caption sits over the very samples it is describing.
        ax.set_xlabel(caption, fontsize=6, color="#64748B", loc="left", labelpad=2)

    ax.set_title(test.get("domainName", test.get("testId", "")), fontsize=8, color="#0F172A", pad=4)
    fig.tight_layout(pad=0.3)
    fig.savefig(path, dpi=DPI, facecolor="white")
    plt.close(fig)
    return bool(pts)


def render_trajectory(seg: dict, path: str) -> bool:
    """
    Target against eye for one Test-mode exercise, X above Y.

    Stacked rather than side by side: the column is 420 px wide, and two panels
    across that leaves each too narrow to read a lag off. Same green/purple as
    the results screen so the two are recognisably the same chart.
    """
    pts = seg.get("points") or []
    fig_w = IMG_W_PX / DPI
    fig, axes = plt.subplots(2, 1, figsize=(fig_w, fig_w * 0.62), dpi=DPI, sharex=True)

    if not pts:
        for ax in axes:
            ax.set_xticks([])
            ax.set_yticks([])
        axes[0].text(0.5, 0.5, "No trajectory recorded", ha="center", va="center",
                     color="#94A3B8", fontsize=9, transform=axes[0].transAxes)
    else:
        t = np.array([p["t"] for p in pts], dtype=float)
        t = t - t.min()
        for ax, axis in zip(axes, ("X", "Y")):
            tgt = np.array([p[f"target{axis}"] for p in pts], dtype=float)
            eye = np.array([p[f"gaze{axis}"] for p in pts], dtype=float)
            ax.plot(t, tgt, color="#4ADE80", linewidth=1.2, label="Target")
            ax.plot(t, eye, color="#A78BFA", linewidth=1.0, label="Eye")
            ax.set_ylabel(f"{axis} (%)", fontsize=6, color="#475569")
            ax.tick_params(labelsize=5, colors="#64748B", length=2)
            ax.grid(True, color="#E2E8F0", linewidth=0.4)
            for spine in ax.spines.values():
                spine.set_edgecolor("#CBD5E1")
                spine.set_linewidth(0.6)
        axes[1].set_xlabel("seconds", fontsize=6, color="#475569")
        axes[0].legend(fontsize=5, loc="upper right", frameon=False, ncol=2)

    rms = seg.get("rmsPct")
    title = seg.get("patternName", "")
    if rms is not None:
        title += f"  ·  {seg.get('pointCount', 0)} pts  ·  RMS {rms:.1f}% of screen"
    axes[0].set_title(title, fontsize=7, color="#0F172A", pad=3)
    fig.tight_layout(pad=0.3)
    fig.savefig(path, dpi=DPI, facecolor="white")
    plt.close(fig)
    return bool(pts)


def style_header(ws, row: int, last_col: int) -> None:
    for c in range(1, last_col + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(vertical="center")


def build_summary(wb: Workbook, data: dict) -> None:
    ws = wb.active
    ws.title = "Summary"

    # Columns A-C are the template's. D onward carry the identity and the
    # conditions, without which "Participant 1" names nobody and an accuracy
    # figure cannot be checked or reproduced.
    headers = [
        "Participants", "Real-time algo Accuracy", "Offline algo Accuracy",
        "Session ID", "Run ID", "Date", "Viewing distance (cm)",
        "Distance measured?", "Real-time error (px)", "Offline error (px)",
        "Quality", "Validation dots", "SD (px)", "Tests scored",
        "Exercise steps", "Mean pursuit RMS (% screen)",
    ]
    ws.append(headers)
    style_header(ws, 1, len(headers))

    for r in data["runs"]:
        acc, off, dist = r["accuracy"], r["offline"], r["distance"]
        scored_tests = sum(1 for t in r["tests"] if t.get("score") is not None)
        ws.append([
            r["rank"],
            acc["angularErrorDeg"],
            off["overallDeg"] if off["overallDeg"] is not None else "not run",
            r["sessionId"],
            r["runId"],
            r["createdAt"][:10],
            dist["configuredCm"],
            "no — configured only" if not dist["measured"] else "yes",
            round(acc["meanErrorPx"], 2),
            round(off["overallPx"], 2) if off["overallPx"] is not None else "",
            acc["quality"],
            acc["validationPoints"],
            round(acc["sdPx"], 2) if acc["sdPx"] is not None else "",
            scored_tests,
            len(r.get("trajectories") or []),
            round(
                sum(s["rmsPct"] for s in r["trajectories"] if s.get("rmsPct") is not None)
                / max(1, sum(1 for s in r["trajectories"] if s.get("rmsPct") is not None)),
                2,
            ) if r.get("trajectories") else "",
        ])

    for row in ws.iter_rows(min_row=2, min_col=2, max_col=3):
        for cell in row:
            if isinstance(cell.value, (int, float)):
                cell.number_format = DEG_FMT

    for col, w in SUMMARY_WIDTHS.items():
        ws.column_dimensions[col].width = w
    for i, w in enumerate([26, 26, 12, 19, 20, 19, 18, 14, 11, 13, 14, 24], start=4):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "B2"

    # What the numbers do and do not mean, on the sheet rather than in a
    # covering email that gets separated from the file.
    row = ws.max_row + 2
    ws.cell(row=row, column=1, value="Notes").font = Font(bold=True)
    totals = data["totals"]
    notes = [
        f"Generated {data['generatedAt'][:19].replace('T', ' ')} · ranked by {data['rankedBy']}.",
        f"{totals['exported']} of {totals['runsWithCalibration']} calibrated runs "
        f"(from {totals['runsExamined']} examined). Median angular error across all: "
        f"{totals['medianAngularErrorDeg']:.2f}°." if totals["medianAngularErrorDeg"] is not None else "",
        f"Offline validation exists for {totals['runsWithOfflineValidation']} run(s); "
        f"the rest show 'not run' rather than a blank that could be read as zero.",
        f"Test-mode exercise steps recorded for {totals.get('runsWithTestModeTrajectories', 0)} run(s); "
        f"each participant sheet lists them below its tests.",
        "Accuracy is angular, not pixel: the same pixel error is twice the angle at half the distance, "
        "so a pixel ranking would reward whoever sat closest to the screen.",
    ] + data["caveats"]
    for n in [x for x in notes if x]:
        row += 1
        ws.cell(row=row, column=1, value=n).font = NOTE_FONT


def build_participant(wb: Workbook, r: dict, img_dir: str, draw_images: bool) -> None:
    ws = wb.create_sheet(f"Participant {r['rank']}")

    ws["B1"] = "Image"
    ws["C1"] = "Link fo the file"  # kept verbatim from the supplied template
    ws["D1"] = "Score"
    ws["E1"] = "Gaze samples"
    ws["F1"] = "Key metrics"
    style_header(ws, 1, 6)

    row = 2
    for t in r["tests"]:
        ws.cell(row=row, column=1, value=t["domainName"]).alignment = Alignment(vertical="center", wrap_text=True)

        rel = os.path.join("report-images", r["runId"], f"{t['testId']}.png")
        abs_png = os.path.join(img_dir, r["runId"], f"{t['testId']}.png")
        if draw_images:
            os.makedirs(os.path.dirname(abs_png), exist_ok=True)
            render_gaze(t, abs_png)
            img = XLImage(abs_png)
            ws.add_image(img, f"B{row}")
            ws.row_dimensions[row].height = img.height * 0.75 + 6
        else:
            ws.cell(row=row, column=2, value="(images skipped)")
            ws.row_dimensions[row].height = 18

        link = ws.cell(row=row, column=3, value=rel)
        link.hyperlink = rel
        link.font = Font(color="2563EB", underline="single")
        link.alignment = Alignment(vertical="center", wrap_text=True)

        ws.cell(row=row, column=4, value=t["score"] if t["score"] is not None else "—").alignment = Alignment(vertical="center")
        ws.cell(row=row, column=5, value=t.get("gazeSampleCount", 0)).alignment = Alignment(vertical="center")
        metrics = t.get("metrics") or {}
        pretty = ", ".join(
            f"{k}={v:.2f}" if isinstance(v, (int, float)) else f"{k}={v}"
            for k, v in list(metrics.items())[:6]
        )
        ws.cell(row=row, column=6, value=pretty or "—").alignment = Alignment(vertical="center", wrap_text=True)
        row += 1

    # The Test-mode exercise steps. They are recorded per session rather than
    # per neurological test, so they are not in `tests` and were missing from
    # the first cut of this report.
    trajectories = r.get("trajectories") or []
    if trajectories:
        row += 1
        band = ws.cell(row=row, column=1, value="Test mode: Target vs Eye tracking")
        band.font = Font(bold=True)
        ws.cell(row=row, column=4, value="Points")
        ws.cell(row=row, column=5, value="RMS error (% of screen)")
        ws.cell(row=row, column=6, value="X / Y split")
        style_header(ws, row, 6)
        row += 1

        for seg in trajectories:
            ws.cell(row=row, column=1, value=seg["patternName"]).alignment = Alignment(vertical="center", wrap_text=True)

            safe = seg["patternName"].lower().replace(" ", "-").replace("/", "-")
            rel = os.path.join("report-images", r["runId"], f"exercise-{safe}.png")
            abs_png = os.path.join(img_dir, r["runId"], f"exercise-{safe}.png")
            if draw_images:
                os.makedirs(os.path.dirname(abs_png), exist_ok=True)
                render_trajectory(seg, abs_png)
                img = XLImage(abs_png)
                ws.add_image(img, f"B{row}")
                ws.row_dimensions[row].height = img.height * 0.75 + 6
            else:
                ws.cell(row=row, column=2, value="(images skipped)")
                ws.row_dimensions[row].height = 18

            link = ws.cell(row=row, column=3, value=rel)
            link.hyperlink = rel
            link.font = Font(color="2563EB", underline="single")
            link.alignment = Alignment(vertical="center", wrap_text=True)

            ws.cell(row=row, column=4, value=seg["pointCount"]).alignment = Alignment(vertical="center")
            ws.cell(row=row, column=5,
                    value=round(seg["rmsPct"], 2) if seg.get("rmsPct") is not None else "—").alignment = Alignment(vertical="center")
            xy = (f"X {seg['rmsXPct']:.1f}% / Y {seg['rmsYPct']:.1f}%"
                  if seg.get("rmsXPct") is not None else "—")
            ws.cell(row=row, column=6, value=xy).alignment = Alignment(vertical="center")
            row += 1

    for col, w in PARTICIPANT_WIDTHS.items():
        ws.column_dimensions[col].width = w
    ws.column_dimensions["D"].width = 8
    ws.column_dimensions["E"].width = 13
    ws.column_dimensions["F"].width = 52

    # Session details below the tests, so the top-left of the sheet stays the
    # shape the template asked for.
    acc, off, dist, cal = r["accuracy"], r["offline"], r["distance"], r["calibration"]
    row += 1
    ws.cell(row=row, column=1, value="Session details").font = Font(bold=True)
    details = [
        ("Run ID", r["runId"]),
        ("Session ID", r["sessionId"]),
        ("Date", r["createdAt"][:19].replace("T", " ")),
        ("Real-time accuracy", f"{acc['angularErrorDeg']:.2f}° ({acc['meanErrorPx']:.1f} px) — {acc['qualityLabel']}"),
        ("Offline accuracy", f"{off['overallDeg']:.2f}° ({off['overallPx']:.1f} px)" if off["overallDeg"] is not None else f"not run (status: {off['status']})"),
        ("Per-dot error (px)", ", ".join(f"{e:.1f}" for e in acc["perPointPx"]) or "—"),
        ("Per-dot spread", f"SD {acc['sdPx']:.1f} px, min {acc['minPx']:.1f}, max {acc['maxPx']:.1f}" if acc["sdPx"] is not None else "—"),
        ("Viewing distance", f"{dist['configuredCm']} cm — configured, not measured ({dist['source']})"),
        ("Face width scale", dist["faceWidthScale"]),
        ("Distance tolerance", dist["headDistanceTolerance"]),
        ("Regression", cal["regressionMethod"]),
        ("Calibration points", cal["calibrationPointsCount"]),
        ("Outlier handling", f"{cal['outlierMethod']} @ {cal['outlierThreshold']}"),
        ("Exercises enabled", cal["enableExercises"]),
        ("Video", r["videoUrl"] or "none"),
    ]
    for label, value in details:
        row += 1
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=9)
        ws.cell(row=row, column=2, value="" if value is None else str(value)).font = Font(size=9)


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--in", dest="src", default="outputs/report-data.json")
    ap.add_argument("--out", dest="out_dir", default="outputs")
    ap.add_argument("--no-images", action="store_true")
    args = ap.parse_args()

    with open(args.src, encoding="utf-8") as f:
        data = json.load(f)

    os.makedirs(args.out_dir, exist_ok=True)
    img_dir = os.path.join(args.out_dir, "report-images")

    wb = Workbook()
    build_summary(wb, data)
    for r in data["runs"]:
        build_participant(wb, r, img_dir, not args.no_images)

    stamp = datetime.now().strftime("%Y-%m-%d_%H%M")
    out = os.path.join(args.out_dir, f"Eye Tracking Report - {stamp}.xlsx")
    wb.save(out)

    n_imgs = (
        sum(len(r["tests"]) + len(r.get("trajectories") or []) for r in data["runs"])
        if not args.no_images else 0
    )
    print(f"Participants: {len(data['runs'])} · images rendered: {n_imgs}")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
